# mypy: disable-error-code="attr-defined"
"""Creates an excel report from error data"""

from collections.abc import Iterable
from dataclasses import dataclass, field
from io import BytesIO
from itertools import chain
from typing import Any, Optional, Union

import polars as pl
from openpyxl import Workbook, utils
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.worksheet import Worksheet
from polars import DataFrame
from polars.exceptions import ColumnNotFoundError

from dve.pipeline.utils import SubmissionStatus


@dataclass
class SummaryItems:
    """Items to go into the Summary sheet"""

    submission_status: SubmissionStatus = field(default_factory=SubmissionStatus)
    """The status of the submission"""
    summary_dict: dict[str, Any] = field(default_factory=dict)
    """Dictionary of items to show in the front sheet key is put into Column B
    and value in column C"""
    row_headings: list[str] = field(default_factory=list)
    """Which errors are expected to show in the summary table"""
    table_columns: list[str] = field(default_factory=list)
    """Names of the tables to show in summary table"""
    partion_key: Optional[str] = None
    """key to split summary items into multiple tables"""
    aggregations: list[pl.Expr] = field(default_factory=lambda: [pl.sum("Count")])  # type: ignore
    """List of aggregations to apply to the grouped up dataframe"""
    additional_columns: Optional[list] = None
    """any additional columns to add to the summary table"""

    def create_summary_sheet(
        self,
        summary: Worksheet,
        aggregates: DataFrame,
        status,
    ) -> Worksheet:
        """Creates a summary sheet for the excel spreadsheet"""
        # Create sheet for report summary

        self._add_submission_info(status, summary)

        try:
            agg_columns = aggregates["Table"].unique().to_list()
        except ColumnNotFoundError:
            agg_columns = []
        tables = self.table_columns or agg_columns
        tables = tables.copy()

        difference = set(agg_columns).difference(tables)
        if difference:
            tables.extend(difference)

        if self.additional_columns:
            tables.extend(self.additional_columns)

        if aggregates.is_empty():
            error_summary = aggregates
        else:
            groups = ["Type", "Table"]

            error_summary = (
                # chaining methods on dataframes seems to confuse mypy
                aggregates.group_by(groups).agg(*self.aggregations)  # type: ignore
            )

        try:
            agg_types = aggregates["Type"].unique().to_list()
        except ColumnNotFoundError:
            agg_types = []

        row_headings = self.row_headings or agg_types
        difference = set(agg_columns).difference(tables)
        if difference:
            tables.extend(difference)

        self._write_table(summary, tables, error_summary, row_headings)

        return summary

    def get_submission_status(self, aggregates: DataFrame) -> str:
        """Returns the status of the submission based on the error data"""
        if self.submission_status.processing_failed:
            return "There was an issue processing the submission. Please contact support."
        if self.submission_status.validation_failed:
            return "File has been rejected"
        if aggregates.is_empty():
            return "File has been accepted, no issues to report"
        failures = aggregates["Type"].unique()
        if "Submission Failure" in failures:
            status = "File has been rejected"
        elif "Warning" in failures:
            status = "File has been accepted, all records accepted with warnings"
        else:
            status = "File has been accepted, no issues to report"
        return status

    def _write_table(
        self, summary: Worksheet, columns: list[str], error_summary: DataFrame, row_headings
    ):
        summary.append(["", "", *columns])
        for error_type in row_headings:
            row: list[Any] = ["", error_type]
            for column in columns:
                if error_summary.is_empty():
                    counts = error_summary
                else:
                    counts = error_summary.filter(  # type: ignore
                        (pl.col("Type") == pl.lit(error_type)) & (pl.col("Table") == pl.lit(column))  # type: ignore # pylint: disable=line-too-long
                    )["Count"]
                if counts.is_empty():
                    row.append(0)
                else:
                    row.append(counts[0])
            summary.append(row)

    def _add_submission_info(self, status: str, summary: Worksheet):
        summary.title = "Summary"
        summary["B2"] = "Data Summary"
        summary.merge_cells("B2:G2")
        summary["B2"].alignment, summary["B2"].font = (
            Alignment(horizontal="center", vertical="center"),
            Font(name="Arial", size=20),
        )
        if "Status" not in self.summary_dict:
            summary.append(["", "Status", status])

        for key, value in self.summary_dict.items():
            summary.append(["", key, str(value)])

        summary.append(["", ""])


class ExcelFormat:
    """Formats error data into an excel file"""

    def __init__(
        self,
        error_details: Union[DataFrame, dict[str, DataFrame]],
        error_aggregates: DataFrame,
        summary_aggregates: Optional[DataFrame] = None,
        overflow=1_000_000,
    ):
        if not isinstance(error_details, dict):
            error_details = {"Error Data": error_details}
        self.error_details = error_details
        """Detailed row by row set of errors"""
        self.aggregates = error_aggregates
        """Aggregated errors usually by error code"""
        self.summary_aggregates = (
            summary_aggregates if summary_aggregates is not None else error_aggregates
        )
        """Aggregates to be used for the front sheet not passed will default to error aggregates"""
        # This is so things can be aggregated to a different degree, for combined reports this
        # is so it can be aggregated to the feed and file level, so numbers from individual files
        # can be displayed.
        self.overflow = overflow
        """Number of errors that will cause an overflow into a new sheet [Default = 1,000,000]"""

    # pylint: disable=too-many-arguments
    def excel_format(
        self,
        summary_items: SummaryItems,
        additional_id: Optional[str] = None,
    ) -> Workbook:
        """Outputs error report to an excel file."""
        # Initialise Workbook
        workbook = Workbook()

        status = summary_items.get_submission_status(self.summary_aggregates)
        active_sheet: Optional[Worksheet] = workbook.active  # type: ignore

        if active_sheet is None:
            active_sheet = workbook.create_sheet()

        summary = summary_items.create_summary_sheet(
            active_sheet,
            self.summary_aggregates,
            status=status,
        )
        self._expand_columns(summary)

        workbook = self.create_error_aggregate_sheet(
            workbook, self.aggregates.iter_rows(), self.aggregates.columns
        )

        for title, df in self.error_details.items():
            workbook.active = workbook.create_sheet(title=title)

            workbook = self.create_error_data_sheets(
                workbook, (df.iter_rows()), df.columns, title=title, additional_id=additional_id
            )
        workbook.active = workbook["Summary"]
        return workbook

    @staticmethod
    def convert_to_bytes(workbook: Workbook) -> bytes:
        """Converts an excel workbook to bytes"""
        # Save workbook to temporary file so it can be passed as a byte stream
        # to write_single_file. This method circumvents the use of named temporary file
        # in the openypxl docs which raises a permission error on Windows
        stream = BytesIO()
        workbook.save(stream)

        return stream.getvalue()

    def create_error_data_sheets(
        self,
        workbook: Workbook,
        invalid_data: Iterable[str],
        headings: list[str],
        title: str = "Error Data",
        suffix: int = 0,
        additional_id: Optional[str] = None,
    ) -> Workbook:
        """Creates a sheet to display error data"""
        # Create sheet for error data
        sheet_title = f"{title}{'_' if suffix else ''}{suffix + 1 if suffix else ''}"
        if suffix == 0:
            error_report: Worksheet = workbook.active  # type: ignore
            error_report.title = sheet_title
        else:
            error_report = workbook.create_sheet(sheet_title)

        headings = self._format_headings(headings)
        if additional_id:
            error_report.append([*headings, additional_id.title()])
        else:
            error_report.append(headings)

        # if not invalid_data:
        #     error_report.append(["No issues to report"])
        #     self._format_error_sheet(error_report)
        #     return workbook

        for row_count, error in enumerate(invalid_data):
            if row_count > self.overflow:
                error_report.append(["Errors continued on next sheet"])
                self._format_error_sheet(error_report)
                return self.create_error_data_sheets(
                    workbook,
                    chain([error], invalid_data),
                    headings,
                    title=title,
                    suffix=suffix + 1,
                )
            row = list(map(str, error))
            try:
                error_report.append(row)
            except IllegalCharacterError:
                for i, item in enumerate(row):
                    if ILLEGAL_CHARACTERS_RE.search(str(item)):
                        row[i] = "Illegal unicode character"
                error_report.append(row)

        self._format_error_sheet(error_report)
        return workbook

    def _format_error_sheet(self, error_report):
        error_report.freeze_panes = "A2"
        error_report.auto_filter.ref = error_report.dimensions

        self._expand_columns(error_report)

    def create_error_aggregate_sheet(
        self, workbook: Workbook, aggregate: list[dict[str, Any]], headings: list[str]
    ):
        """Creates a sheet aggregating errors together to give a more granular overview"""
        # Create sheet for error summary info
        error_report: Worksheet = workbook.create_sheet("Error Summary")

        headings = self._format_headings(headings)

        error_report.append(headings)

        for item in aggregate:
            error_report.append(list(item))

        error_report.freeze_panes = "A2"
        error_report.auto_filter.ref = error_report.dimensions

        self._expand_columns(error_report)

        return workbook

    def _expand_columns(self, worksheet):
        for column_cells in worksheet.columns:
            length = min(80, max(self._text_length(cell.value) for cell in column_cells))  # type: ignore # pylint: disable=line-too-long

            worksheet.column_dimensions[utils.get_column_letter(column_cells[0].column)].width = (
                length + 5
            )

    @staticmethod
    def _text_length(value):
        return 0 if value is None else len(str(value))

    @staticmethod
    def _format_headings(headings: list[str]) -> list[str]:
        # TODO - ideally this would be config driven to allow customisation.
        _renames = {
            "Table": "Group",
            "Data Item": "Data Item Submission Name",
            "Error": "Errors and Warnings",
        }
        headings = [heading.title() if heading[0].islower() else heading for heading in headings]
        headings = [heading.replace("_", " ") for heading in headings]
        headings = [_renames.get(heading, heading) for heading in headings]
        return headings
